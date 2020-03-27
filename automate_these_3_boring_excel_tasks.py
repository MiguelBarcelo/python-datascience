#########################
# Automate These 3 (Boring!!) Excel Tasks with Python!
# By Nik Piepenbreier 
# towardsdatascience.com
# Mar 25
# https://towardsdatascience.com/automate-these-3-boring-excel-tasks-with-python-666b4ded101b
#########################

# 1. Combining Multiple Excel Files
# You might find yourself with a number of Excel workbooks (such as monthly sales reports). 
# One day, you’re asked to calculate the total number of sales across all these reports. 
# We can easily combine Excel workbooks with Python, using the Pandas library.
# You can install Pandas using either pip or conda:
# > pip install pandas
# > conda install pandas
# For this tutorial, let’s load in three separate Excel workbooks that are linked to below:
# File 1: https://github.com/datagy/mediumdata/raw/master/january.xlsx
# File 2: https://github.com/datagy/mediumdata/raw/master/february.xlsx
# File 3: https://github.com/datagy/mediumdata/raw/master/march.xlsx
# We can see that the data doesn’t really start until Row 4, so we’ll need Pandas to import sheets beginning at that row. In the code below, we’ll make use of the read_excel and append functions.

# Section 1
# we imported pandas, created a list with all the URLs, and generated a blank dataframe called combined.
import pandas as pd
files = ['https://github.com/datagy/mediumdata/raw/master/january.xlsx', 'https://github.com/datagy/mediumdata/raw/master/february.xlsx', 'https://github.com/datagy/mediumdata/raw/master/march.xlsx']
combined = pd.DataFrame()

# Section 2
# we looped through each of the URLs in files to read each file into a dataframe (“df”), skipping the first three rows, and append it to the combined dataframe.
for file in files:
	df = pd.read_excel(file, skiprows = 3)
	combined = combined.append(df, ignore_index = True)

# Section 3
# we write generate a new Excel file called combined.xlsx containing our merged Excel workbooks!
combined.to_excel('combined.xlsx')



# 2. Getting Values from Various Workbooks
# Let’s take a look at another example! Say we needed to get only the total for Toronto 
# from each sales report and collect them in a list. We know that the total is stored in cell F5 
# in every workbook. If you’re following along, this package works if you have your files stored locally.
# Download the files with the links above and save them to your machine.
# For this example, we’ll use a different library called openpyxl. 
# You can install it with either pip or conda using the code below:
# > pip install openpyxl
# > onda install openpyxl

# Section 1
# Generated a list (“files”) that contains links to all our files. In Windows we can Shift+right-click and use Copy as Path to get its path.
# You will likely want to convert the string to a raw string, by prefixing it with ‘r’.
# We also generated an empty list to store our values.
import openpyxl

files 	= [] # include paths to your files here
values 	= []

# Section 2
# Looped over the files using openpyxl.
# The .load_workbook() method loads a file.
# We use [‘Sheet1’] and [‘F5’] to reference both sheet names and cell references in workbook and worksheet objects.
# Finally, we use the .value attribute to extract the cell’s value and append it to the values list.
for file in files:
	wb = openpyxl.load_workbook(file)
	sheet = wb['Sheet1']
	value = sheet['F5'].value
	values.append(value)



# 3. Applying Formulas across Workbooks
# Let’s take a look at one final example! In each of the Excel workbooks, we have totals across the rows, 
# but not a grand total for sales. Again, we could open each workbook and add in a formula, 
# or we can use Python to do this for us!
# We’ll be making use of openpyxl again. If you need to install it, check out the instructions above. 
# The links to download the files are also available above.

# Section 1
import openpyxl

files = [] # insert paths to files here

# Section 2
# we again fill a list to the files. The for-loop, opens each file and assigns “Sheet1” to a variable sheet.
# We then assign the string ‘=SUM(F5:F8)’ to cell F9 and use the .style attribute to assign 
# the currency style directly to the cell. More cell styles can be found in the official documentation.
for file in files:
	wb = openpyxl.load_workbook(file)
	sheet = wb['Sheet1']
	sheet['F9'] = '=SUM(F5:F8)'
	sheet['F9'].style = 'Currency'
	wb.save(file)

